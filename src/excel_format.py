"""Общее оформление листов Excel (v1 и v2)."""

from __future__ import annotations

import re
from datetime import date, datetime
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from src.date_utils import parse_date_column
from src.excel_sanitize import sanitize_dataframe
from src.tab_number import format_tab_number_columns, tab_number_column_labels

# Ключи config.columns / snapshot_columns, которые являются датами
_DATE_COLUMN_KEYS: tuple[str, ...] = (
    "report_date",
    "work_start_date",
    "deal_created_date",
    "team_added_date",
)


def cell_text_width(value: Any) -> int:
    """Ширина для автоподбора: для многострочных — длина самой длинной строки."""
    if value is None:
        return 0
    text: str = str(value)
    if "\n" in text:
        return max((len(line) for line in text.splitlines()), default=0)
    return len(text)


def date_column_labels(config: dict[str, Any]) -> list[str]:
    """Excel-заголовки колонок с датами (из snapshot_columns и columns)."""
    labels: list[str] = []
    seen: set[str] = set()
    snap: dict[str, Any] = dict(config.get("output", {}).get("snapshot_columns") or {})
    columns: dict[str, Any] = dict(config.get("columns") or {})
    for key in _DATE_COLUMN_KEYS:
        label: str | None = None
        if key in snap and snap[key]:
            label = str(snap[key])
        elif key in columns and columns[key]:
            label = str(columns[key])
        if label and label not in seen:
            seen.add(label)
            labels.append(label)
    return labels


def coerce_date_columns(frame: pd.DataFrame, config: dict[str, Any]) -> pd.DataFrame:
    """Приводит известные колонки дат к datetime (для записи как даты Excel)."""
    if frame.empty:
        return frame
    labels: set[str] = set(date_column_labels(config))
    if not labels:
        return frame
    out: pd.DataFrame = frame.copy()
    for col_name in out.columns:
        header: str = str(col_name)
        if header not in labels:
            continue
        out[col_name] = parse_date_column(out[col_name], config, header)
    return out


def prepare_excel_frame(frame: pd.DataFrame, config: dict[str, Any]) -> pd.DataFrame:
    """Санитизация, даты и нормализация табельных номеров перед записью в Excel."""
    prepared: pd.DataFrame = coerce_date_columns(frame, config)
    prepared = format_tab_number_columns(prepared, config)
    return sanitize_dataframe(prepared)


def _parse_last_col(value: Any) -> int:
    """
    Последний закреплённый столбец → номер (1=A).
    0 / null / '' — столбцы не закреплять.
    Допускается число или буква колонки («C»).
    """
    if value is None or value == "":
        return 0
    if isinstance(value, bool):
        return 0
    if isinstance(value, int):
        return max(0, int(value))
    if isinstance(value, float):
        return max(0, int(value))
    text: str = str(value).strip().upper()
    if not text or text in {"0", "NONE", "NULL"}:
        return 0
    if text.isdigit():
        return max(0, int(text))
    if re.fullmatch(r"[A-Z]+", text):
        return int(column_index_from_string(text))
    return 0


def freeze_panes_from_last(*, last_row: int, last_col: int) -> str | None:
    """
    Строит адрес ячейки freeze_panes Excel.

    last_row / last_col — последние закреплённые строка и столбец (1-based; col 0 = нет).
    Excel ставит freeze на первую незакреплённую ячейку.
    Пример: last_row=1, last_col=0 → A2; last_row=3, last_col=3 → D4.
    """
    row: int = max(0, int(last_row))
    col: int = max(0, int(last_col))
    if row <= 0 and col <= 0:
        return None
    first_free_row: int = row + 1 if row > 0 else 1
    first_free_col: int = col + 1 if col > 0 else 1
    return f"{get_column_letter(first_free_col)}{first_free_row}"


def resolve_freeze_panes(
    config: dict[str, Any],
    sheet_key: str | None = None,
    *,
    default: str | None = "A2",
) -> str | None:
    """
    Закрепление для листа из output.sheet_freeze.

    Формат блока листа: { "last_row": 1, "last_col": 0 } — последние закреплённые.
    Fallback: sheet_freeze.default → excel_format.freeze_panes → default.
    """
    out_cfg: dict[str, Any] = config.get("output") or {}
    freeze_map: dict[str, Any] = dict(out_cfg.get("sheet_freeze") or {})

    block: Any = None
    if sheet_key and sheet_key in freeze_map:
        block = freeze_map[sheet_key]
    elif "default" in freeze_map:
        block = freeze_map["default"]

    if isinstance(block, dict) and ("last_row" in block or "last_col" in block):
        last_row: int = int(block.get("last_row") or 0)
        last_col: int = _parse_last_col(block.get("last_col"))
        return freeze_panes_from_last(last_row=last_row, last_col=last_col)

    if isinstance(block, str) and block.strip():
        # Совместимость: строка вида "A2" / "D4"
        return block.strip()

    fmt_cfg: dict[str, Any] = out_cfg.get("excel_format") or {}
    legacy: Any = fmt_cfg.get("freeze_panes")
    if isinstance(legacy, str) and legacy.strip():
        return legacy.strip()
    return default


def format_sheet(
    ws: Any,
    config: dict[str, Any],
    *,
    sheet_key: str | None = None,
) -> None:
    """Применяет автофильтр, закрепление, ширину и раскраску min/max."""
    if ws.max_row < 1 or ws.max_column < 1:
        return

    fmt_cfg: dict[str, Any] = config["output"]["excel_format"]
    labels: dict[str, str] = config["output"]["column_labels"]
    theme: str = config.get("excel_theme", "green_red")
    colors: dict[str, str] = fmt_cfg.get("colors", {"min": "C6EFCE", "max": "FFC7CE"})

    green_fill: PatternFill = PatternFill(
        start_color=colors["min"], end_color=colors["min"], fill_type="solid"
    )
    red_fill: PatternFill = PatternFill(
        start_color=colors["max"], end_color=colors["max"], fill_type="solid"
    )

    freeze: str | None = resolve_freeze_panes(config, sheet_key)
    if freeze:
        ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions

    min_marker: str = labels.get("min_header_marker", "Мин")
    max_marker: str = labels.get("max_header_marker", "Макс")
    headers: list[str] = [cell.value for cell in ws[1]]
    min_cols: list[int] = []
    max_cols: list[int] = []

    for idx, header in enumerate(headers, start=1):
        if header and min_marker in str(header):
            min_cols.append(idx)
        if header and max_marker in str(header):
            max_cols.append(idx)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in min_cols:
            cell = row[col_idx - 1]
            if theme == "green_red" and isinstance(cell.value, (int, float)):
                cell.fill = green_fill
        for col_idx in max_cols:
            cell = row[col_idx - 1]
            if theme == "green_red" and isinstance(cell.value, (int, float)):
                cell.fill = red_fill

    min_width: int = int(fmt_cfg.get("min_column_width", 12))
    max_width: int = int(fmt_cfg.get("max_column_width", 45))
    sample_rows: int = int(fmt_cfg.get("sample_rows_for_width", 200))

    for col_idx in range(1, ws.max_column + 1):
        letter: str = get_column_letter(col_idx)
        width: int = min_width
        for row_idx in range(1, min(ws.max_row, sample_rows) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            width = max(width, min(cell_text_width(value) + 2, max_width))
        ws.column_dimensions[letter].width = width

    header_font: Font = Font(bold=True)
    header_align: Alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_align

    float_fmt: str = fmt_cfg.get("float_format", "0.00")
    int_fmt: str = fmt_cfg.get("int_format", "0")
    date_fmt: str = str(fmt_cfg.get("date_format", "DD.MM.YYYY"))
    tab_cols: set[str] = set(tab_number_column_labels(config))
    date_cols: set[str] = set(date_column_labels(config))
    headers_map: dict[int, str] = {idx: str(header or "") for idx, header in enumerate(headers, start=1)}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx, cell in enumerate(row, start=1):
            header_name: str = headers_map.get(col_idx, "")
            if isinstance(cell.value, (datetime, date)):
                cell.number_format = date_fmt
            elif isinstance(cell.value, float):
                cell.number_format = float_fmt
            elif isinstance(cell.value, int):
                cell.number_format = int_fmt
            if header_name in tab_cols and cell.value is not None:
                cell.number_format = "@"
            elif header_name in date_cols and isinstance(cell.value, (datetime, date)):
                cell.number_format = date_fmt
            wrap: bool = bool(cell.alignment.wrap_text) if cell.alignment else False
            horizontal: str | None = cell.alignment.horizontal if cell.alignment else None
            cell.alignment = Alignment(
                horizontal=horizontal,
                vertical="center",
                wrap_text=wrap,
            )


def format_multiline_columns(
    ws: Any,
    config: dict[str, Any],
    header_markers: list[str],
) -> None:
    """Перенос по словам и высота строк для колонок с многострочным текстом."""
    if ws.max_row < 1 or not header_markers:
        return

    headers: list[Any] = [cell.value for cell in ws[1]]
    fmt_cfg: dict[str, Any] = config.get("output", {}).get("excel_format", {})
    max_width: int = int(fmt_cfg.get("hotspots_column_width", 55))
    wrap_align: Alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")

    for marker in header_markers:
        col_idx: int | None = None
        for idx, header in enumerate(headers, start=1):
            if header and marker in str(header):
                col_idx = idx
                break
        if col_idx is None:
            continue

        letter: str = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = max_width
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = wrap_align
            text: str = str(cell.value or "")
            lines: int = max(1, text.count("\n") + 1) if text and text != "—" else 1
            ws.row_dimensions[row_idx].height = max(15.0, min(15.0 * lines, 120.0))
