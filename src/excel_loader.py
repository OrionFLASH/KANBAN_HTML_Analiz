"""Параллельная загрузка Excel-файлов Kanban."""

from __future__ import annotations

import logging
import re
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

logger: logging.Logger = logging.getLogger("kanban.excel_loader")

REQUIRED_COLUMNS: list[str] = [
    "Дата отчета",
    "ID ПрПр",
    "Группа продукта",
    "Продукт",
    "Дата начала работы",
    "Текущий статус",
    "Количество дней на текущей стадии",
    "Дата создания сделки",
    "Стадия сделки",
    "Количество дней с создания сделки",
    "ТБ",
    "_Изменение условий",
    "_Ввод данных",
    "ЕФС флаг",
    "Метка",
]

_EMPTY_STAGE_VALUES: set[str] = {"", "-", "nan", "None"}


def _detect_category(filename: str) -> str:
    """Определяет категорию файла по имени (не используется в аналитике)."""
    if "К ПРОДАЖЕ" in filename:
        return "К ПРОДАЖЕ"
    if "В РАБОТЕ" in filename:
        return "В РАБОТЕ"
    return "UNKNOWN"


def _resolve_table_range(
    file_path: Path,
    sheet_name: str,
    table_name: str,
) -> str | None:
    """Возвращает диапазон именованной таблицы Excel или None."""
    wb = load_workbook(file_path, read_only=False, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        for tbl in ws.tables.values():
            if tbl.name == table_name:
                return tbl.ref
        return None
    finally:
        wb.close()


def _read_table_range(file_path: Path, sheet_name: str, cell_range: str) -> pd.DataFrame:
    """Читает именованную таблицу Excel по диапазону ref."""
    match = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", cell_range.replace("$", ""))
    if not match:
        raise ValueError(f"Некорректный диапазон таблицы: {cell_range}")

    start_row: int = int(match.group(2))
    end_row: int = int(match.group(4))
    return pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        engine="openpyxl",
        header=start_row - 1,
        nrows=end_row - start_row,
        na_values=[""],
    )


def read_single_file(args: tuple[str, dict[str, Any]]) -> pd.DataFrame:
    """Читает один Excel-файл (для ProcessPoolExecutor)."""
    file_path_str, config = args
    file_path: Path = Path(file_path_str)
    sheet_name: str = config["sheet_name"]
    table_name: str = config.get("excel_table_name", "Base")
    use_auto: bool = bool(config.get("excel_table_auto", True))

    df: pd.DataFrame
    if use_auto:
        cell_range: str | None = _resolve_table_range(file_path, sheet_name, table_name)
        if cell_range:
            df = _read_table_range(file_path, sheet_name, cell_range)
        else:
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                engine="openpyxl",
                na_values=[""],
            )
    else:
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            engine="openpyxl",
            na_values=[""],
        )

    _validate_columns(df, file_path)
    df = _normalize_types(df)
    df["source_file"] = file_path.name
    df["source_category"] = _detect_category(file_path.name)
    return df


def _validate_columns(df: pd.DataFrame, file_path: Path) -> None:
    """Проверяет наличие обязательных колонок."""
    missing: list[str] = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(f"{file_path.name}: отсутствуют колонки: {missing}")


def _normalize_types(df: pd.DataFrame) -> pd.DataFrame:
    """Приводит ключевые колонки к нужным типам."""
    result: pd.DataFrame = df.copy()
    date_cols: list[str] = ["Дата отчета", "Дата начала работы"]
    for col in date_cols:
        if col in result.columns:
            result[col] = pd.to_datetime(result[col], errors="coerce")

    if "Дата создания сделки" in result.columns:
        result["Дата создания сделки"] = pd.to_datetime(
            result["Дата создания сделки"], errors="coerce"
        )

    int_cols: list[str] = [
        "Количество дней на текущей стадии",
        "_Изменение условий",
        "_Ввод данных",
        "ЕФС флаг",
    ]
    for col in int_cols:
        if col in result.columns:
            result[col] = pd.to_numeric(result[col], errors="coerce")

    if "Количество дней с создания сделки" in result.columns:
        result["Количество дней с создания сделки"] = pd.to_numeric(
            result["Количество дней с создания сделки"], errors="coerce"
        )

    for col in ["Текущий статус", "Стадия сделки", "Группа продукта", "Продукт", "ТБ", "ID ПрПр"]:
        if col in result.columns:
            result[col] = result[col].astype(str).str.strip()

    if "Стадия сделки" in result.columns:
        result["Стадия сделки"] = result["Стадия сделки"].replace("nan", "")

    return result


def load_all_files(config: dict[str, Any], input_dir: Path, filenames: list[str]) -> pd.DataFrame:
    """Параллельно загружает все файлы и объединяет в один DataFrame."""
    paths: list[Path] = [input_dir / name for name in filenames]
    for path in paths:
        if not path.exists():
            raise FileNotFoundError(f"Входной файл не найден: {path}")

    workers: int = int(config.get("parallel_workers", 4))
    args_list: list[tuple[str, dict[str, Any]]] = [(str(p), config) for p in paths]

    frames: list[pd.DataFrame] = []
    if workers == 1 or len(paths) == 1:
        for args in args_list:
            frames.append(read_single_file(args))
    else:
        with ProcessPoolExecutor(max_workers=workers) as executor:
            futures = {executor.submit(read_single_file, args): args[0] for args in args_list}
            for future in as_completed(futures):
                path_str: str = futures[future]
                try:
                    frames.append(future.result())
                    logger.info("Загружен файл: %s", Path(path_str).name)
                except Exception as exc:
                    logger.error("Ошибка загрузки %s: %s", path_str, exc)
                    raise

    if not frames:
        raise ValueError("Не удалось загрузить ни одного файла")

    combined: pd.DataFrame = pd.concat(frames, ignore_index=True)
    logger.info("Объединено строк: %d из %d файлов", len(combined), len(frames))
    return combined
