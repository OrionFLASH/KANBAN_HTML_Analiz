"""Экспорт Excel v2 на несколько листов (форматирование как в основном pipeline)."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from src.excel_format import (
    format_multiline_columns,
    format_sheet,
    freeze_panes_from_last,
    prepare_excel_frame,
    resolve_freeze_panes,
)
from src.excel_sanitize import sanitize_sheet_name
from src.export_overflow import (
    build_csv_redirect_sheet,
    export_overflow_csv_sheets,
    split_sheets_by_row_limit,
)
from src.percentile_stats import percentile_label
from src.settings import percentile_display_value
from src.v2.duration_matrix import DurationMatrixResult

logger: logging.Logger = logging.getLogger("kanban.excel_v2.exporter")

# Колонки с переводами строк — ширина и wrap как у «Топ зон» в run.py
MULTILINE_BY_SHEET: dict[str, list[str]] = {
    "managers": ["Группа + Продукт", "Почта Альфа", "Почта Сигма"],
    "leads": [
        "TN Лидера лида",
        "ФИО Лидера лида",
        "Почта Альфа Лидера лида",
        "Почта Сигма Лидера лида",
        "Роль Лидера лида",
        "ТБ Лидера лида",
        "TN Лидера сделки",
        "ФИО Лидера сделки",
        "Почта Альфа Лидера сделки",
        "Почта Сигма Лидера сделки",
        "Роль Лидера сделки",
        "ТБ Лидера сделки",
        "Клиент",
    ],
    "violations": ["Клиент", "Почта Альфа", "Почта Сигма"],
}


def _write_dataframe_block(
    ws: Any,
    frame: pd.DataFrame,
    start_row: int,
    *,
    thousands_format: str | None = None,
) -> int:
    """Пишет DataFrame с заголовком начиная со start_row; возвращает следующую свободную строку."""
    if frame is None or frame.empty:
        return start_row
    n_header_rows: int = 0
    for r_idx, row in enumerate(dataframe_to_rows(frame, index=False, header=True)):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=start_row + r_idx, column=c_idx, value=value)
            # Заголовок (r_idx==0) не форматируем; числа — с разделителем разрядов
            if (
                thousands_format
                and r_idx > 0
                and isinstance(value, (int, float))
                and not isinstance(value, bool)
            ):
                cell.number_format = thousands_format
        n_header_rows = r_idx + 1
    return start_row + n_header_rows


def _style_block_header(ws: Any, header_row: int, n_cols: int) -> None:
    """Жирный заголовок таблицы на указанной строке."""
    header_font: Font = Font(bold=True)
    header_align: Alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx in range(1, max(n_cols, 1) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.alignment = header_align


def _autosize_columns(ws: Any, max_col: int, max_row: int, config: dict[str, Any]) -> None:
    """Простая ширина колонок для листа «Статистика»."""
    fmt_cfg: dict[str, Any] = config.get("output", {}).get("excel_format", {})
    min_width: int = int(fmt_cfg.get("min_column_width", 12))
    max_width: int = int(fmt_cfg.get("max_column_width", 45))
    for col_idx in range(1, max_col + 1):
        letter: str = get_column_letter(col_idx)
        width: int = min_width
        for row_idx in range(1, min(max_row, 200) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            # С разделителем разрядов число визуально длиннее
            text: str = f"{value:,}".replace(",", " ") if isinstance(value, (int, float)) and not isinstance(value, bool) else str(value)
            width = max(width, min(len(text) + 2, max_width))
        ws.column_dimensions[letter].width = width


def _write_statistics_sheet(
    ws: Any,
    config: dict[str, Any],
    funnel_frame: pd.DataFrame | None,
    outlier_summary: pd.DataFrame | None,
) -> None:
    """
    Лист «Статистика»: воронка фильтров + свод выбросов.
    Не использует format_sheet (несколько блоков) — своё оформление.
    Числа — формат с разделителем разрядов (# ##0).
    """
    row: int = 1
    title_font: Font = Font(bold=True, size=12)
    fmt_cfg: dict[str, Any] = config.get("output", {}).get("excel_format", {})
    # Пробел как разделитель тысяч (Excel: # ##0)
    thousands_format: str = str(fmt_cfg.get("thousands_format", "# ##0"))

    ws.cell(row=row, column=1, value="Воронка отсечения по фильтрам (строки Kanban / уникальные лиды)")
    ws.cell(row=row, column=1).font = title_font
    row += 1
    funnel_header_row: int = row
    if funnel_frame is not None and not funnel_frame.empty:
        row = _write_dataframe_block(
            ws, funnel_frame, row, thousands_format=thousands_format
        )
        _style_block_header(ws, funnel_header_row, len(funnel_frame.columns))
        # Автофильтр только на таблицу воронки
        end_col: str = get_column_letter(len(funnel_frame.columns))
        end_row: int = row - 1
        ws.auto_filter.ref = f"A{funnel_header_row}:{end_col}{end_row}"
        freeze_stats: str | None = resolve_freeze_panes(config, "statistics", default=None)
        if freeze_stats:
            ws.freeze_panes = freeze_stats
        else:
            ws.freeze_panes = f"A{funnel_header_row + 1}"
    else:
        ws.cell(row=row, column=1, value="(нет шагов фильтров)")
        row += 1
        freeze_stats = resolve_freeze_panes(config, "statistics", default=None)
        if freeze_stats:
            ws.freeze_panes = freeze_stats

    row += 1
    ws.cell(
        row=row,
        column=1,
        value="Свод отсечения выбросов (сумма по группам; детали по ТБ/группе/продукту/стадии — на листе «Нормативы»)",
    )
    ws.cell(row=row, column=1).font = title_font
    row += 1
    summary_header_row: int = row
    if outlier_summary is not None and not outlier_summary.empty:
        row = _write_dataframe_block(
            ws, outlier_summary, row, thousands_format=thousands_format
        )
        _style_block_header(ws, summary_header_row, len(outlier_summary.columns))
    else:
        ws.cell(row=row, column=1, value="(нет данных по выбросам)")
        row += 1

    _autosize_columns(ws, max(ws.max_column or 1, 1), max(ws.max_row or 1, 1), config)


def _write_duration_matrix_sheet(
    ws: Any,
    matrix: DurationMatrixResult,
    config: dict[str, Any],
) -> None:
    """
    Лист «Распределение сроков»:
    стр.1 шапка, стр.2 итоги по дням, стр.3 номера колонок + автофильтр,
    далее продукты; колонки: группа | продукт | P… | Всего | дни;
    закрепление до «Всего»; пунктирные границы; жирная рамка на дне порога.
    """
    labels: dict[str, Any] = config.get("output", {}).get("column_labels") or {}
    mtx_cfg: dict[str, Any] = config.get("output", {}).get("duration_matrix") or {}
    pg_label: str = str(labels.get("product_group", "Группа продукта"))
    pr_label: str = str(labels.get("product", "Продукт"))
    total_label: str = str(mtx_cfg.get("total_column_label", "Всего"))
    day_width: float = float(mtx_cfg.get("day_column_width", 4.5))
    pct_width: float = float(mtx_cfg.get("percentile_column_width", 8))
    row_height: float = float(mtx_cfg.get("row_height", 28))
    header_row_height: float = float(mtx_cfg.get("header_row_height", 22))
    filter_row_height: float = float(mtx_cfg.get("filter_row_height", 16))
    label_widths: dict[str, float] = mtx_cfg.get("label_column_widths") or {}
    int_fmt: str = str(
        config.get("output", {}).get("excel_format", {}).get("int_format", "0")
    )
    pale_yellow: str = str(mtx_cfg.get("header_fill", "FFF2CC"))
    faint_color: str = str(mtx_cfg.get("filter_row_font_color", "D9D9D9"))
    grid_color: str = str(mtx_cfg.get("grid_border_color", "BFBFBF"))
    highlight_color: str = str(mtx_cfg.get("threshold_border_color", "C65911"))

    yellow_fill: PatternFill = PatternFill(
        start_color=pale_yellow, end_color=pale_yellow, fill_type="solid"
    )
    center: Alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap: Alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    header_font: Font = Font(bold=True)
    product_font: Font = Font(bold=True)
    faint_font: Font = Font(color=faint_color, size=8)
    counts_font_size: int = int(mtx_cfg.get("counts_font_size", 14))
    counts_font: Font = Font(size=counts_font_size)
    totals_font: Font = Font(bold=True, size=counts_font_size)

    dash_side: Side = Side(style="dashed", color=grid_color)
    dashed_border: Border = Border(
        left=dash_side, right=dash_side, top=dash_side, bottom=dash_side
    )
    thick_side: Side = Side(style="medium", color=highlight_color)
    threshold_border: Border = Border(
        left=thick_side, right=thick_side, top=thick_side, bottom=thick_side
    )

    color_cfg: dict[str, Any] = mtx_cfg.get("color_scale") or {}
    start_color: str = str(color_cfg.get("start", "63BE7B"))
    mid_color: str = str(color_cfg.get("mid", "FFEB84"))
    end_color: str = str(color_cfg.get("end", "F8696B"))

    if matrix.empty or not matrix.rows:
        ws.cell(row=1, column=1, value="Нет данных для матрицы сроков")
        return

    pct_list: list[float] = list(matrix.percentile_list)
    n_pct: int = len(pct_list)
    # A=группа, B=продукт, C… = процентили, затем Всего, затем дни
    total_col: int = 2 + n_pct + 1
    days_start_col: int = total_col + 1
    last_col: int = total_col + len(matrix.day_columns)

    pct_headers: list[str] = [
        f"P{percentile_display_value(p)}" for p in pct_list
    ]
    headers: list[Any] = [pg_label, pr_label, *pct_headers, total_label, *matrix.day_columns]
    for col_idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=value)
        cell.font = header_font
        cell.alignment = center
        cell.fill = yellow_fill

    # Строка 2 — горизонтальные итоги по дням (процентили пустые)
    total_pg = ws.cell(row=2, column=1, value=total_label)
    total_pg.font = header_font
    total_pg.alignment = center
    total_pg.fill = yellow_fill
    for col_idx in range(2, total_col):
        empty_cell = ws.cell(row=2, column=col_idx, value="")
        empty_cell.alignment = center
        empty_cell.fill = yellow_fill
    grand_cell = ws.cell(
        row=2,
        column=total_col,
        value=matrix.grand_total if matrix.grand_total else None,
    )
    grand_cell.font = totals_font
    grand_cell.alignment = center
    grand_cell.fill = yellow_fill
    if matrix.grand_total:
        grand_cell.number_format = int_fmt
    for day_offset, day in enumerate(matrix.day_columns):
        day_sum: int = int(matrix.day_totals.get(day, 0))
        cell = ws.cell(
            row=2,
            column=days_start_col + day_offset,
            value=day_sum if day_sum else None,
        )
        cell.font = totals_font
        cell.alignment = center
        cell.fill = yellow_fill
        if day_sum:
            cell.number_format = int_fmt

    # Строка 3 — еле видимые номера колонок + автофильтр
    filter_row: int = 3
    for col_idx in range(1, last_col + 1):
        cell = ws.cell(row=filter_row, column=col_idx, value=col_idx)
        cell.font = faint_font
        cell.alignment = center

    data_start: int = 4
    exc_p: float = float(matrix.exceedance_percentile)
    highlight_cells: list[tuple[int, int]] = []

    for row_offset, (pg, pr, total, counts, pcts) in enumerate(matrix.rows, start=data_start):
        pg_cell = ws.cell(row=row_offset, column=1, value=pg)
        pg_cell.alignment = center
        pr_cell = ws.cell(row=row_offset, column=2, value=pr)
        pr_cell.alignment = left_wrap
        pr_cell.font = product_font
        pr_cell.fill = yellow_fill

        for pct_i, p in enumerate(pct_list):
            pct_days: int | None = pcts.get(float(p))
            pct_cell = ws.cell(
                row=row_offset,
                column=3 + pct_i,
                value=pct_days if pct_days is not None else None,
            )
            pct_cell.alignment = center
            pct_cell.font = counts_font
            if pct_days is not None:
                pct_cell.number_format = int_fmt

        total_cell = ws.cell(
            row=row_offset, column=total_col, value=total if total else None
        )
        total_cell.alignment = center
        total_cell.font = counts_font
        if total:
            total_cell.number_format = int_fmt

        threshold_day: int | None = pcts.get(exc_p)
        for day_offset, day in enumerate(matrix.day_columns):
            cnt: int | None = counts.get(day)
            col_idx = days_start_col + day_offset
            if cnt:
                cell = ws.cell(row=row_offset, column=col_idx, value=int(cnt))
                cell.alignment = center
                cell.font = counts_font
                cell.number_format = int_fmt
            if threshold_day is not None and int(day) == int(threshold_day):
                highlight_cells.append((row_offset, col_idx))

    last_row: int = data_start - 1 + len(matrix.rows)

    # Пунктирные границы на всю таблицу, затем жирная рамка порога
    for row_idx in range(1, last_row + 1):
        for col_idx in range(1, last_col + 1):
            ws.cell(row=row_idx, column=col_idx).border = dashed_border
    for row_idx, col_idx in highlight_cells:
        ws.cell(row=row_idx, column=col_idx).border = threshold_border

    if last_row >= data_start and last_col >= days_start_col:
        data_range: str = (
            f"{get_column_letter(days_start_col)}{data_start}:"
            f"{get_column_letter(last_col)}{last_row}"
        )
        ws.conditional_formatting.add(
            data_range,
            ColorScaleRule(
                start_type="min",
                start_color=start_color,
                mid_type="percentile",
                mid_value=50,
                mid_color=mid_color,
                end_type="max",
                end_color=end_color,
            ),
        )
        total_letter: str = get_column_letter(total_col)
        ws.conditional_formatting.add(
            f"{total_letter}{data_start}:{total_letter}{last_row}",
            ColorScaleRule(
                start_type="min",
                start_color=start_color,
                mid_type="percentile",
                mid_value=50,
                mid_color=mid_color,
                end_type="max",
                end_color=end_color,
            ),
        )

    # Закрепление: последний столбец — «Всего»
    freeze: str | None = freeze_panes_from_last(last_row=3, last_col=total_col)
    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A{filter_row}:{get_column_letter(last_col)}{last_row}"
    ws.column_dimensions["A"].width = float(label_widths.get("A", 28))
    ws.column_dimensions["B"].width = float(label_widths.get("B", 36))
    for pct_i in range(n_pct):
        ws.column_dimensions[get_column_letter(3 + pct_i)].width = pct_width
    ws.column_dimensions[get_column_letter(total_col)].width = float(
        label_widths.get("total", label_widths.get("C", 10))
    )
    for col_idx in range(days_start_col, last_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = day_width
    ws.row_dimensions[1].height = header_row_height
    ws.row_dimensions[2].height = header_row_height
    ws.row_dimensions[filter_row].height = filter_row_height
    for row_idx in range(data_start, last_row + 1):
        ws.row_dimensions[row_idx].height = row_height

    logger.info(
        "Матрица сроков: лист записан, колонок процентилей=%s, порог %s, "
        "выделено ячеек=%s, freeze=%s",
        n_pct,
        percentile_label(exc_p),
        len(highlight_cells),
        freeze,
    )


def export_excel_v2(
    path: Path,
    sheets: dict[str, pd.DataFrame],
    config: dict[str, Any],
    *,
    funnel_frame: pd.DataFrame | None = None,
    outlier_summary: pd.DataFrame | None = None,
    duration_matrix: DurationMatrixResult | None = None,
) -> tuple[Path, list[Path]]:
    """
    Записывает листы в Excel; листы > excel_max_rows_per_sheet — в CSV (;).
    Возвращает (путь xlsx, список путей csv).

    «Нормативы» — обычная таблица групп (+ колонки отсечения выбросов по строке).
    «Статистика» — воронка фильтров и свод выбросов.
    «Распределение сроков» — матрица группа/продукт × дни.
    """
    path.parent.mkdir(parents=True, exist_ok=True)

    sheet_names: dict[str, str] = dict(config.get("output", {}).get("sheets") or {})
    max_len: int = int(config.get("output", {}).get("excel_max_sheet_name_length", 31))
    engine: str = str(config.get("excel", {}).get("engine", "openpyxl"))

    excel_sheets, csv_sheets = split_sheets_by_row_limit(sheets, config)
    csv_paths: list[Path] = export_overflow_csv_sheets(path, csv_sheets, sheet_names, config)

    used_sheet_names: set[str] = set()
    prepared: dict[str, pd.DataFrame] = {}
    sheet_key_by_title: dict[str, str] = {}

    # Порядок листов: нормативы, статистика, матрица сроков, остальные
    preferred_order: list[str] = [
        "norms",
        "statistics",
        "duration_matrix",
        "leads",
        "managers",
        "violations",
    ]
    ordered_keys: list[str] = [k for k in preferred_order if k in excel_sheets]
    ordered_keys.extend([k for k in excel_sheets if k not in ordered_keys])

    want_duration: bool = duration_matrix is not None and not duration_matrix.empty

    if not excel_sheets and csv_paths and not want_duration:
        redirect_title: str = sanitize_sheet_name("Экспорт CSV", used_sheet_names, max_len)
        prepared[redirect_title] = build_csv_redirect_sheet(csv_paths)
        sheet_key_by_title[redirect_title] = "_csv_redirect"
    else:
        for key in ordered_keys:
            frame = excel_sheets[key]
            title: str = sanitize_sheet_name(sheet_names.get(key, key), used_sheet_names, max_len)
            sheet_key_by_title[title] = key
            if key in {"statistics", "duration_matrix"}:
                # Плейсхолдер — содержимое пишется отдельно
                prepared[title] = pd.DataFrame({"_": []})
            else:
                prepared[title] = prepare_excel_frame(frame, config)

        # Если в sheets не передали statistics, но есть funnel — всё равно создаём лист
        if "statistics" not in excel_sheets and (
            (funnel_frame is not None and not funnel_frame.empty)
            or (outlier_summary is not None and not outlier_summary.empty)
        ):
            title = sanitize_sheet_name(
                sheet_names.get("statistics", "Статистика"), used_sheet_names, max_len
            )
            sheet_key_by_title[title] = "statistics"
            prepared[title] = pd.DataFrame({"_": []})

        if want_duration and "duration_matrix" not in excel_sheets:
            title = sanitize_sheet_name(
                sheet_names.get("duration_matrix", "Распределение сроков"),
                used_sheet_names,
                max_len,
            )
            sheet_key_by_title[title] = "duration_matrix"
            # Вставить после статистики, если возможно
            prepared[title] = pd.DataFrame({"_": []})

    with pd.ExcelWriter(path, engine=engine) as writer:
        for title, frame in prepared.items():
            key: str = sheet_key_by_title.get(title, "")
            if key == "statistics":
                pd.DataFrame({"_": []}).to_excel(writer, sheet_name=title, index=False)
                ws = writer.book[title]
                if ws.max_row >= 1:
                    ws.delete_rows(1, ws.max_row)
                _write_statistics_sheet(ws, config, funnel_frame, outlier_summary)
            elif key == "duration_matrix":
                pd.DataFrame({"_": []}).to_excel(writer, sheet_name=title, index=False)
                ws = writer.book[title]
                if ws.max_row >= 1:
                    ws.delete_rows(1, ws.max_row)
                if duration_matrix is not None:
                    _write_duration_matrix_sheet(ws, duration_matrix, config)
            elif frame.empty:
                pd.DataFrame({"Нет данных": []}).to_excel(writer, sheet_name=title, index=False)
            else:
                frame.to_excel(writer, sheet_name=title, index=False)

        wb = writer.book
        for title in wb.sheetnames:
            if title.startswith("_"):
                continue
            ws = wb[title]
            sheet_key: str = sheet_key_by_title.get(title, "")
            if sheet_key in {"statistics", "duration_matrix"}:
                # Уже оформлены своими writers
                continue
            format_sheet(ws, config, sheet_key=sheet_key)
            markers: list[str] = MULTILINE_BY_SHEET.get(sheet_key, [])
            if markers:
                format_multiline_columns(ws, config, markers)

        wb.save(path)

    if csv_paths:
        logger.info(
            "Excel v2: %s (%s листов в xlsx, %s в CSV)",
            path,
            len(prepared),
            len(csv_paths),
        )
    else:
        logger.info("Excel v2 сохранён: %s (%s листов)", path, len(prepared))
    return path, csv_paths
