"""Экспорт результатов в форматированный Excel."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.excel_sanitize import sanitize_dataframe, sanitize_sheet_name
from src.export_overflow import (
    build_csv_redirect_sheet,
    export_overflow_csv_sheets,
    split_sheets_by_row_limit,
)
from src.manager_analytics import manager_analytics_to_excel_frame
from src.settings import col
from src.statistics_config import build_statistics_export_mapping, filter_and_order_statistics_frame
from src.tab_number import format_tab_number_columns, tab_number_column_labels

logger: logging.Logger = logging.getLogger("kanban.excel_exporter")


def _build_export_mapping(config: dict[str, Any]) -> dict[str, str]:
    """Строит mapping колонок DataFrame → заголовки Excel из config."""
    labels: dict[str, str] = config["output"]["column_labels"]
    mapping: dict[str, str] = {
        col(config, "product_group"): labels.get("product_group", "ГРУППА"),
        col(config, "product"): labels.get("product", "ПРОДУКТ"),
        col(config, "tb"): labels.get("tb", "ТБ"),
        "current_status": labels.get("current_status", "Текущий статус"),
        "deal_stage": labels.get("deal_stage", "Стадия сделки"),
        "stage_key": labels.get("stage_key", "Ключ стадии"),
        "analysis_level": labels.get("analysis_level", "Уровень анализа"),
    }
    mapping.update(build_statistics_export_mapping(config))
    return mapping


def _prepare_excel_frame(frame: pd.DataFrame, config: dict[str, Any]) -> pd.DataFrame:
    """Санитизация и нормализация табельных номеров перед записью в Excel."""
    return sanitize_dataframe(format_tab_number_columns(frame, config))


def _rename_columns_for_export(df: pd.DataFrame, config: dict[str, Any]) -> pd.DataFrame:
    """Переименовывает и упорядочивает колонки для читаемого Excel."""
    trimmed: pd.DataFrame = filter_and_order_statistics_frame(df, config)
    return trimmed.rename(columns=_build_export_mapping(config))


def _cell_text_width(value: Any) -> int:
    """Ширина для автоподбора: для многострочных — длина самой длинной строки."""
    if value is None:
        return 0
    text: str = str(value)
    if "\n" in text:
        return max((len(line) for line in text.splitlines()), default=0)
    return len(text)


def _format_sheet(ws, config: dict[str, Any]) -> None:
    """Применяет автофильтр, закрепление, ширину и раскраску min/max."""
    if ws.max_row < 1 or ws.max_column < 1:
        return

    fmt_cfg: dict[str, Any] = config["output"]["excel_format"]
    labels: dict[str, str] = config["output"]["column_labels"]
    theme: str = config.get("excel_theme", "green_red")
    colors: dict[str, str] = fmt_cfg.get("colors", {"min": "C6EFCE", "max": "FFC7CE"})

    green_fill: PatternFill = PatternFill(
        start_color=colors["min"], end_color=colors["min"], fill_type="solid"
    )
    red_fill: PatternFill = PatternFill(
        start_color=colors["max"], end_color=colors["max"], fill_type="solid"
    )

    ws.freeze_panes = fmt_cfg.get("freeze_panes", "A2")
    ws.auto_filter.ref = ws.dimensions

    min_marker: str = labels.get("min_header_marker", "Мин")
    max_marker: str = labels.get("max_header_marker", "Макс")
    headers: list[str] = [cell.value for cell in ws[1]]
    min_cols: list[int] = []
    max_cols: list[int] = []

    for idx, header in enumerate(headers, start=1):
        if header and min_marker in str(header):
            min_cols.append(idx)
        if header and max_marker in str(header):
            max_cols.append(idx)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in min_cols:
            cell = row[col_idx - 1]
            if theme == "green_red" and isinstance(cell.value, (int, float)):
                cell.fill = green_fill
        for col_idx in max_cols:
            cell = row[col_idx - 1]
            if theme == "green_red" and isinstance(cell.value, (int, float)):
                cell.fill = red_fill

    min_width: int = int(fmt_cfg.get("min_column_width", 12))
    max_width: int = int(fmt_cfg.get("max_column_width", 45))
    sample_rows: int = int(fmt_cfg.get("sample_rows_for_width", 200))

    for col_idx in range(1, ws.max_column + 1):
        letter: str = get_column_letter(col_idx)
        width: int = min_width
        for row_idx in range(1, min(ws.max_row, sample_rows) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            width = max(width, min(_cell_text_width(value) + 2, max_width))
        ws.column_dimensions[letter].width = width

    header_font: Font = Font(bold=True)
    header_align: Alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align: Alignment = Alignment(vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_align

    float_fmt: str = fmt_cfg.get("float_format", "0.00")
    int_fmt: str = fmt_cfg.get("int_format", "0")
    tab_cols: set[str] = set(tab_number_column_labels(config))
    headers_map: dict[int, str] = {idx: str(header or "") for idx, header in enumerate(headers, start=1)}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx, cell in enumerate(row, start=1):
            if isinstance(cell.value, float):
                cell.number_format = float_fmt
            elif isinstance(cell.value, int):
                cell.number_format = int_fmt
            header_name: str = headers_map.get(col_idx, "")
            if header_name in tab_cols and cell.value is not None:
                cell.number_format = "@"
            wrap: bool = bool(cell.alignment.wrap_text) if cell.alignment else False
            horizontal: str | None = cell.alignment.horizontal if cell.alignment else None
            cell.alignment = Alignment(
                horizontal=horizontal,
                vertical="center",
                wrap_text=wrap,
            )


def _format_multiline_columns(
    ws,
    config: dict[str, Any],
    header_markers: list[str],
) -> None:
    """Перенос по словам и высота строк для колонок с многострочным текстом."""
    if ws.max_row < 1 or not header_markers:
        return

    headers: list[Any] = [cell.value for cell in ws[1]]
    fmt_cfg: dict[str, Any] = config.get("output", {}).get("excel_format", {})
    max_width: int = int(fmt_cfg.get("hotspots_column_width", 55))
    wrap_align: Alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")

    for marker in header_markers:
        col_idx: int | None = None
        for idx, header in enumerate(headers, start=1):
            if header and marker in str(header):
                col_idx = idx
                break
        if col_idx is None:
            continue

        letter: str = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = max_width
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = wrap_align
            text: str = str(cell.value or "")
            lines: int = max(1, text.count("\n") + 1) if text and text != "—" else 1
            ws.row_dimensions[row_idx].height = max(15.0, min(15.0 * lines, 120.0))


def _format_managers_hotspots_column(ws, config: dict[str, Any]) -> None:
    """Перенос по словам и высота строк для колонки «Топ зон превышения»."""
    _format_multiline_columns(ws, config, ["Топ зон"])


def export_excel(
    stats: dict[str, Any],
    output_path: Path,
    config: dict[str, Any],
    manager_payload: dict[str, Any] | None = None,
) -> None:
    """Записывает Excel с листами статистики и менеджеров (без «Графики»)."""
    out_cfg: dict[str, Any] = config["output"]
    sheet_names: dict[str, str] = out_cfg["excel_sheets"]
    max_len: int = int(out_cfg.get("excel_max_sheet_name_length", 31))

    output_path.parent.mkdir(parents=True, exist_ok=True)

    used_sheet_names: set[str] = set()
    sheets: dict[str, pd.DataFrame] = {
        sanitize_sheet_name(sheet_names["summary"], used_sheet_names, max_len): _prepare_excel_frame(
            _rename_columns_for_export(stats["by_tb"], config),
            config,
        ),
        sanitize_sheet_name(sheet_names["overall"], used_sheet_names, max_len): _prepare_excel_frame(
            _rename_columns_for_export(stats["overall"], config),
            config,
        ),
    }
    for tb_name, tb_df in stats.get("tb_sheets", {}).items():
        safe_name: str = sanitize_sheet_name(str(tb_name), used_sheet_names, max_len)
        sheets[safe_name] = _prepare_excel_frame(_rename_columns_for_export(tb_df, config), config)

    managers_sheet: str | None = None
    if manager_payload:
        managers_sheet = sanitize_sheet_name(
            sheet_names.get("managers", "Менеджеры"), used_sheet_names, max_len
        )
        mgr_frame: pd.DataFrame = manager_analytics_to_excel_frame(manager_payload, config)
        if not mgr_frame.empty:
            sheets[managers_sheet] = _prepare_excel_frame(mgr_frame, config)
        else:
            managers_sheet = None

    # Листы > excel_max_rows_per_sheet — в CSV (;), не во вкладку Excel
    excel_sheets, csv_sheets = split_sheets_by_row_limit(sheets, config)
    csv_titles: dict[str, str] = {name: name for name in csv_sheets}
    csv_paths: list[Path] = export_overflow_csv_sheets(
        output_path, csv_sheets, csv_titles, config
    )

    if not excel_sheets and csv_paths:
        redirect_used: set[str] = set()
        redirect_title: str = sanitize_sheet_name("Экспорт CSV", redirect_used, max_len)
        excel_sheets = {redirect_title: build_csv_redirect_sheet(csv_paths)}

    with pd.ExcelWriter(output_path, engine=config["excel"].get("engine", "openpyxl")) as writer:
        for sheet_name, frame in excel_sheets.items():
            if frame.empty:
                continue
            frame.to_excel(writer, sheet_name=sheet_name, index=False)

        wb = writer.book

        for sheet_name in wb.sheetnames:
            if sheet_name.startswith("_"):
                continue
            _format_sheet(wb[sheet_name], config)

        if managers_sheet and managers_sheet in wb.sheetnames:
            _format_managers_hotspots_column(wb[managers_sheet], config)

        wb.save(output_path)

    if csv_paths:
        logger.info("Excel сохранён: %s (%s листов в xlsx, %s CSV)", output_path, len(excel_sheets), len(csv_paths))
    else:
        logger.info("Excel сохранён: %s", output_path)
