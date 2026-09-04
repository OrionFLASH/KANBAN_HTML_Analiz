"""Тесты матрицы сроков: группа/продукт × дни."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.v2.duration_matrix import build_duration_matrix
from src.v2.exporter import export_excel_v2


def _base_config(*, sort_mode: str = "alpha_days") -> dict:
    return {
        "columns": {
            "product_group": "Группа продукта",
            "product": "Продукт",
            "days_on_stage": "Дней на стадии",
        },
        "percentiles": [20, 50, 80],
        "exceedance": {"percentile": 50},
        "excel": {"engine": "openpyxl"},
        "output": {
            "sheets": {
                "norms": "Нормативы",
                "duration_matrix": "Распределение сроков",
                "leads": "Уникальные ID",
            },
            "column_labels": {
                "product_group": "Группа продукта",
                "product": "Продукт",
                "min_header_marker": "Мин",
                "max_header_marker": "Макс",
            },
            "duration_matrix": {
                "enabled": True,
                "sort_mode": sort_mode,
                "total_column_label": "Всего",
                "day_column_width": 4.5,
                "percentile_column_width": 8,
                "row_height": 28,
                "header_row_height": 22,
                "filter_row_height": 16,
                "counts_font_size": 14,
                "header_fill": "FFF2CC",
                "filter_row_font_color": "D9D9D9",
                "grid_border_color": "BFBFBF",
                "threshold_border_color": "C65911",
                "max_day_span": 3000,
                "color_scale": {
                    "start": "63BE7B",
                    "mid": "FFEB84",
                    "end": "F8696B",
                },
            },
            "sheet_freeze": {
                "duration_matrix": {"last_row": 3, "last_col": 6},
                "default": {"last_row": 1, "last_col": 0},
            },
            "excel_format": {
                "freeze_panes": "A2",
                "min_column_width": 10,
                "max_column_width": 40,
                "sample_rows_for_width": 50,
                "float_format": "0.00",
                "int_format": "0",
                "colors": {"min": "C6EFCE", "max": "FFC7CE"},
            },
            "excel_max_rows_per_sheet": 900000,
            "csv_overflow": {"enabled": False},
        },
        "excel_theme": "green_red",
    }


def _sample_snap() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "product_group": ["Бета", "Альфа", "Альфа", "Альфа", "Бета"],
            "product": ["Z", "B", "A", "A", "Z"],
            "_days_on_stage": [5.2, 3.0, 3.0, 10.0, 5.0],
        }
    )


def test_build_duration_matrix_alpha_days() -> None:
    result = build_duration_matrix(_sample_snap(), _base_config(sort_mode="alpha_days"))
    assert not result.empty
    assert result.sort_mode == "alpha_days"
    assert result.day_columns == [3, 5, 10]
    assert result.day_totals == {3: 2, 5: 2, 10: 1}
    assert result.grand_total == 5
    assert result.percentile_list == [20.0, 50.0, 80.0]
    assert result.exceedance_percentile == 50.0
    assert [(r[0], r[1]) for r in result.rows] == [
        ("Альфа", "A"),
        ("Альфа", "B"),
        ("Бета", "Z"),
    ]
    alpha_a = result.rows[0]
    assert alpha_a[2] == 2
    assert alpha_a[3][3] == 1
    assert alpha_a[3][10] == 1
    # Два лида: 3 и 10 → P50 = 3 (нижние 50% = 1 лид → max=3)
    assert alpha_a[4][50.0] == 3
    assert result.rows[2][3][5] == 2


def test_build_duration_matrix_by_volume() -> None:
    result = build_duration_matrix(_sample_snap(), _base_config(sort_mode="by_volume"))
    assert result.sort_mode == "by_volume"
    # Дни 3 и 5 по 2 лида — слева меньший день при равенстве; 10 с 1 — правее
    assert result.day_columns == [3, 5, 10]
    # Строки: Бета/Z и Альфа/A по 2 лида — алфавит Альфа раньше; Альфа/B = 1
    assert [(r[0], r[1], r[2]) for r in result.rows] == [
        ("Альфа", "A", 2),
        ("Бета", "Z", 2),
        ("Альфа", "B", 1),
    ]


def test_duration_matrix_sheet_format(tmp_path: Path) -> None:
    config = _base_config(sort_mode="by_volume")
    snap = pd.DataFrame(
        {
            "product_group": ["G1", "G1", "G2"],
            "product": ["P1", "P1", "P2"],
            "_days_on_stage": [1, 3, 1],
        }
    )
    matrix = build_duration_matrix(snap, config)
    # День 1: 2 лида, день 3: 1 → слева день 1
    assert matrix.day_columns == [1, 3]
    # P1: дни 1 и 3 → P50 = 1
    p1 = next(r for r in matrix.rows if r[1] == "P1")
    assert p1[4][50.0] == 1

    path = tmp_path / "matrix.xlsx"
    export_excel_v2(
        path,
        {
            "norms": pd.DataFrame({"ТБ": ["x"]}),
            "duration_matrix": pd.DataFrame(),
            "leads": pd.DataFrame({"ID": [1]}),
        },
        config,
        duration_matrix=matrix,
    )
    wb = load_workbook(path)
    ws = wb["Распределение сроков"]
    # Группа | Продукт | P20 | P50 | P80 | Всего | дни → freeze после Всего (кол. 6) = G4
    assert ws.freeze_panes == "G4"
    assert ws.auto_filter.ref is not None
    assert ws.auto_filter.ref.startswith("A3:")
    assert ws.cell(1, 1).value == "Группа продукта"
    assert ws.cell(1, 2).value == "Продукт"
    assert ws.cell(1, 3).value == "P20"
    assert ws.cell(1, 4).value == "P50"
    assert ws.cell(1, 5).value == "P80"
    assert ws.cell(1, 6).value == "Всего"
    assert ws.cell(1, 7).value == 1
    assert ws.cell(1, 8).value == 3
    assert ws.cell(2, 1).value == "Всего"
    assert ws.cell(2, 6).value == 3
    # строка номеров колонок
    assert ws.cell(3, 6).value == 6
    assert ws.cell(3, 7).value == 7
    # данные с 4-й строки; сверху продукт с большим total
    assert ws.cell(4, 2).value == "P1"
    assert ws.cell(4, 2).alignment.horizontal == "left"
    assert ws.cell(4, 2).font.bold is True
    assert ws.cell(1, 1).fill.fgColor.rgb in {"00FFF2CC", "FFF2CC"}
    assert ws.cell(4, 4).value == 1  # P50
    assert ws.cell(4, 6).value == 2  # Всего
    assert ws.cell(4, 6).font.size == 14
    assert ws.row_dimensions[4].height == 28
    # пунктир на обычной ячейке
    assert ws.cell(4, 1).border.left.style == "dashed"
    # выделение порога P50=1 → колонка дня 1 (col 7)
    assert ws.cell(4, 7).border.left.style == "medium"
    assert ws.cell(4, 7).border.left.color.rgb in {"00C65911", "C65911"}
