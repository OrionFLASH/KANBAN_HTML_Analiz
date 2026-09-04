"""Тесты формата дат в Excel."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.excel_format import (
    coerce_date_columns,
    date_column_labels,
    format_sheet,
    prepare_excel_frame,
)
from src.excel_sanitize import sanitize_cell_value


def _config() -> dict:
    return {
        "columns": {
            "report_date": "Дата отчета",
            "work_start_date": "Дата начала работы",
            "deal_created_date": "Дата создания сделки",
        },
        "dates": {
            "dayfirst": True,
            "formats": ["%d.%m.%Y", "%Y-%m-%d"],
            "empty_values": ["", "-", "nan", "none", "nat"],
        },
        "output": {
            "snapshot_columns": {
                "work_start_date": "Дата начала работы",
                "deal_created_date": "Дата создания сделки",
            },
            "column_labels": {
                "min_header_marker": "Мин",
                "max_header_marker": "Макс",
            },
            "excel_format": {
                "date_format": "DD.MM.YYYY",
                "float_format": "0.00",
                "int_format": "0",
                "min_column_width": 12,
                "max_column_width": 45,
                "sample_rows_for_width": 200,
                "colors": {"min": "C6EFCE", "max": "FFC7CE"},
            },
        },
        "team_files": {"output_columns": {}},
        "excel_theme": "green_red",
    }


def test_date_column_labels() -> None:
    labels: list[str] = date_column_labels(_config())
    assert "Дата начала работы" in labels
    assert "Дата создания сделки" in labels
    assert "Дата отчета" in labels


def test_coerce_and_sanitize_keeps_datetime() -> None:
    cfg: dict = _config()
    frame: pd.DataFrame = pd.DataFrame(
        {
            "Дата начала работы": ["01.09.2026", "2026-08-15", None],
            "Клиент": ["A", "B", "C"],
        }
    )
    prepared: pd.DataFrame = prepare_excel_frame(frame, cfg)
    assert isinstance(prepared.loc[0, "Дата начала работы"], datetime)
    assert prepared.loc[0, "Дата начала работы"].day == 1
    assert prepared.loc[0, "Дата начала работы"].month == 9
    assert prepared.loc[1, "Дата начала работы"].day == 15
    assert prepared.loc[2, "Дата начала работы"] is None


def test_sanitize_preserves_python_date() -> None:
    assert isinstance(sanitize_cell_value(date(2026, 9, 1)), date)
    assert isinstance(sanitize_cell_value(datetime(2026, 9, 1, 12, 0)), datetime)


def test_format_sheet_applies_date_number_format(tmp_path: Path) -> None:
    cfg: dict = _config()
    frame: pd.DataFrame = coerce_date_columns(
        pd.DataFrame(
            {
                "ID": ["L1"],
                "Дата начала работы": ["01.09.2026"],
                "Дата создания сделки": [pd.Timestamp("2026-08-31")],
            }
        ),
        cfg,
    )
    prepared: pd.DataFrame = prepare_excel_frame(frame, cfg)
    path: Path = tmp_path / "dates.xlsx"
    prepared.to_excel(path, index=False)
    wb = load_workbook(path)
    ws = wb.active
    format_sheet(ws, cfg, sheet_key="leads")
    assert isinstance(ws.cell(2, 2).value, datetime)
    assert ws.cell(2, 2).number_format == "DD.MM.YYYY"
    assert isinstance(ws.cell(2, 3).value, datetime)
    assert ws.cell(2, 3).number_format == "DD.MM.YYYY"
