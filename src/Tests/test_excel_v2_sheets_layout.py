"""Тест: Нормативы — обычная таблица; Статистика — воронка отдельно."""

from __future__ import annotations

from pathlib import Path

import pandas as pd

from src.v2.exporter import export_excel_v2


def test_norms_is_plain_table_statistics_has_funnel(tmp_path: Path) -> None:
    config: dict = {
        "excel": {"engine": "openpyxl"},
        "output": {
            "sheets": {
                "norms": "Нормативы",
                "statistics": "Статистика",
                "leads": "Уникальные ID",
            },
            "sheet_freeze": {
                "norms": {"last_row": 1, "last_col": 0},
                "default": {"last_row": 1, "last_col": 0},
            },
            "column_labels": {"min_header_marker": "Мин", "max_header_marker": "Макс"},
            "excel_format": {
                "freeze_panes": "A2",
                "min_column_width": 10,
                "max_column_width": 40,
                "sample_rows_for_width": 50,
                "float_format": "0.00",
                "int_format": "0",
                "thousands_format": "# ##0",
                "colors": {"min": "C6EFCE", "max": "FFC7CE"},
            },
            "excel_max_rows_per_sheet": 900000,
            "csv_overflow": {"enabled": False},
        },
        "excel_theme": "green_red",
    }
    norms = pd.DataFrame(
        {
            "ТБ": ["ТБ1"],
            "Группа продукта": ["G"],
            "Продукт": ["P"],
            "Стадия работы с лидом": ["В работе"],
            "Число лидов": [3],
            "До отсечения": [4],
            "После отсечения": [3],
            "Отсечено (всего)": [1],
        }
    )
    funnel = pd.DataFrame(
        {
            "Этап": ["Загрузка", "Фильтр: efs"],
            "До (строк)": [100, 100],
            "После (строк)": [100, 80],
            "Отсечено строк": [0, 20],
            "До (лидов)": [50, 50],
            "После (лидов)": [50, 40],
            "Отсечено лидов": [0, 10],
        }
    )
    outlier_summary = pd.DataFrame(
        {"Показатель": ["Отсечено (всего)"], "Значение": [1]}
    )
    path: Path = tmp_path / "report.xlsx"
    export_excel_v2(
        path,
        {
            "norms": norms,
            "statistics": pd.DataFrame(),
            "leads": pd.DataFrame({"ID": [1]}),
        },
        config,
        funnel_frame=funnel,
        outlier_summary=outlier_summary,
    )
    assert path.exists()
    xl = pd.ExcelFile(path)
    assert "Нормативы" in xl.sheet_names
    assert "Статистика" in xl.sheet_names

    norms_df = pd.read_excel(path, sheet_name="Нормативы")
    # Шапка сразу в первой строке — без блока «Воронка…»
    assert list(norms_df.columns)[0] == "ТБ"
    assert "Отсечено (всего)" in norms_df.columns
    assert "Воронка" not in str(norms_df.columns)

    stats_df = pd.read_excel(path, sheet_name="Статистика", header=None)
    # Первая ячейка — заголовок воронки
    assert "Воронка" in str(stats_df.iloc[0, 0])

    from openpyxl import load_workbook

    wb = load_workbook(path)
    norms_ws = wb["Нормативы"]
    assert norms_ws.freeze_panes == "A2"
    assert norms_ws.auto_filter.ref is not None

    stats_ws = wb["Статистика"]
    # Числовая ячейка воронки (после заголовка блока и шапки таблицы) — формат с разрядами
    found_thousands: bool = False
    for row in stats_ws.iter_rows(min_row=1, max_row=20, max_col=10):
        for cell in row:
            if isinstance(cell.value, int) and cell.value >= 10:
                assert "# ##0" in str(cell.number_format).replace(",", " ") or cell.number_format == "# ##0"
                found_thousands = True
                break
        if found_thousands:
            break
    assert found_thousands
