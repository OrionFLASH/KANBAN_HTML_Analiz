"""Листы сводной матрицы и графиков в Excel."""

from __future__ import annotations

import logging
from typing import Any

import pandas as pd
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.marker import Marker
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from src.settings import analysis_row_key, is_group_only_analysis
from src.visualization_data import build_pivot_matrix, indicator_keys, series_chart_points

logger: logging.Logger = logging.getLogger("kanban.pivot_excel")

HIDDEN_SHEETS: tuple[str, ...] = ("_pivot_flat", "_chart_src")


def _metric_labels(config: dict[str, Any]) -> dict[str, str]:
    """Человекочитаемые названия метрик."""
    return {
        "days_on_stage": "Дни на стадии",
        "days_since_deal": "Дни с создания сделки",
    }


def _indicator_labels(config: dict[str, Any]) -> dict[str, str]:
    """Человекочитаемые названия показателей."""
    labels: dict[str, str] = {"min": "Мин", "max": "Макс"}
    for p in config.get("percentiles", [20, 50, 80]):
        key: str = f"p{int(p)}" if float(p).is_integer() else f"p{p}"
        labels[key] = f"П{int(p)}" if float(p).is_integer() else f"П{p}"
    return labels


def write_hidden_pivot_flat(ws: Worksheet, pivot_flat: list[dict[str, Any]]) -> None:
    """Записывает длинную таблицу для lookup."""
    headers: list[str] = [
        "tb",
        "product_group",
        "product",
        "stage_key",
        "metric",
        "indicator",
        "value",
    ]
    ws.append(headers)
    for row in pivot_flat:
        ws.append([row.get(h) for h in headers])
    ws.sheet_state = "hidden"


def write_hidden_chart_source(ws: Worksheet, distribution_series: list[dict[str, Any]]) -> None:
    """Плоская таблица точек для построения графиков."""
    headers: list[str] = [
        "series_id",
        "tb",
        "product",
        "stage_key",
        "metric",
        "lead_index",
        "days",
    ]
    ws.append(headers)
    for idx, series in enumerate(distribution_series):
        for point in series_chart_points(series):
            ws.append(
                [
                    idx,
                    series.get("tb"),
                    series.get("product"),
                    series.get("stage_key"),
                    series.get("metric"),
                    point.get("lead_index"),
                    point.get("days"),
                ]
            )
    ws.sheet_state = "hidden"


def _write_matrix_table(
    ws: Worksheet,
    matrix: dict[str, Any],
    start_row: int,
    start_col: int,
    row_header: str = "Продукт",
) -> tuple[int, int]:
    """Пишет блок матрицы строка × стадия."""
    stages: list[str] = list(matrix["stages"])
    row_labels: list[str] = list(matrix.get("rows") or matrix.get("products") or [])
    values: dict[str, dict[str, int | None]] = matrix["values"]

    header_row: int = start_row
    ws.cell(row=header_row, column=start_col, value=row_header)
    for col_idx, stage in enumerate(stages, start=start_col + 1):
        cell = ws.cell(row=header_row, column=col_idx, value=stage)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_offset, row_label in enumerate(row_labels, start=1):
        row_idx: int = header_row + row_offset
        ws.cell(row=row_idx, column=start_col, value=row_label)
        for col_offset, stage in enumerate(stages, start=1):
            value = values.get(row_label, {}).get(stage)
            cell = ws.cell(row=row_idx, column=start_col + col_offset, value=value)
            cell.number_format = "0"

    last_row: int = header_row + len(row_labels)
    last_col: int = start_col + len(stages)
    return last_row, last_col


def write_matrix_sheet(
    ws: Worksheet,
    pivot_flat: list[dict[str, Any]],
    distribution_series: list[dict[str, Any]],
    config: dict[str, Any],
) -> None:
    """Лист «Матрица» с выпадающими фильтрами и таблицей."""
    dash_cfg: dict[str, Any] = config.get("dashboard", {})
    all_tb_label: str = str(dash_cfg.get("all_tb_label", "__ALL__"))
    all_tb_display: str = str(dash_cfg.get("all_tb_display", "ВСЕ ТБ"))
    default_tb: str = str(dash_cfg.get("default_tb", all_tb_label))
    default_metric: str = str(dash_cfg.get("default_metric", "days_on_stage"))
    default_indicator: str = str(dash_cfg.get("default_indicator", "p80"))

    tb_values: list[str] = sorted({str(row["tb"]) for row in pivot_flat})
    if all_tb_label not in tb_values:
        tb_values.insert(0, all_tb_label)
    tb_values_display: list[str] = [
        all_tb_display if value == all_tb_label else value for value in tb_values
    ]

    metric_labels: dict[str, str] = _metric_labels(config)
    indicator_labels: dict[str, str] = _indicator_labels(config)
    metrics: list[str] = list(config["aggregation"].get("metrics", ["days_on_stage", "days_since_deal"]))
    indicators: list[str] = indicator_keys(config)

    ws["A1"] = "ТБ"
    ws["B1"] = all_tb_display if default_tb == all_tb_label else default_tb
    ws["A2"] = "Показатель"
    ws["B2"] = indicator_labels.get(default_indicator, default_indicator)
    ws["A3"] = "Метрика"
    ws["B3"] = metric_labels.get(default_metric, default_metric)

    for label_cell in ("A1", "A2", "A3"):
        ws[label_cell].font = Font(bold=True)

    # Списки для выпадающих фильтров на листе _lists (скрытый фрагмент в конце матрицы)
    list_col: int = 20
    ws.cell(row=1, column=list_col, value="tb_list")
    for idx, tb in enumerate(tb_values_display, start=2):
        ws.cell(row=idx, column=list_col, value=tb)
    tb_last: int = 1 + len(tb_values)

    ws.cell(row=1, column=list_col + 1, value="indicator_list")
    indicator_keys_internal: list[str] = indicators
    for idx, key in enumerate(indicator_keys_internal, start=2):
        ws.cell(row=idx, column=list_col + 1, value=indicator_labels.get(key, key))
    ind_last: int = 1 + len(indicator_keys_internal)

    ws.cell(row=1, column=list_col + 2, value="metric_list")
    for idx, key in enumerate(metrics, start=2):
        ws.cell(row=idx, column=list_col + 2, value=metric_labels.get(key, key))
    met_last: int = 1 + len(metrics)

    tb_letter: str = get_column_letter(list_col)
    ind_letter: str = get_column_letter(list_col + 1)
    met_letter: str = get_column_letter(list_col + 2)

    dv_tb = DataValidation(type="list", formula1=f"=${tb_letter}$2:${tb_letter}${tb_last}", allow_blank=False)
    dv_ind = DataValidation(type="list", formula1=f"=${ind_letter}$2:${ind_letter}${ind_last}", allow_blank=False)
    dv_met = DataValidation(type="list", formula1=f"=${met_letter}$2:${met_letter}${met_last}", allow_blank=False)
    ws.add_data_validation(dv_tb)
    ws.add_data_validation(dv_ind)
    ws.add_data_validation(dv_met)
    dv_tb.add(ws["B1"])
    dv_ind.add(ws["B2"])
    dv_met.add(ws["B3"])

    matrix: dict[str, Any] = build_pivot_matrix(pivot_flat, default_tb, default_metric, default_indicator, config)
    row_header: str = "Группа" if is_group_only_analysis(config) else "Продукт"
    last_row, last_col = _write_matrix_table(
        ws, matrix, start_row=5, start_col=1, row_header=row_header
    )

    ws.cell(row=4, column=1, value="Свод: строка × стадия (значение — выбранный показатель, целые дни)")
    ws["A4"].font = Font(italic=True)

    if last_row >= 6 and last_col >= 2:
        color_range: str = f"B6:{get_column_letter(last_col)}{last_row}"
        ws.conditional_formatting.add(
            color_range,
            ColorScaleRule(
                start_type="min",
                start_color="63BE7B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="F8696B",
            ),
        )

    ws.freeze_panes = "B6"
    ws.column_dimensions["A"].width = 42
    for col_idx in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    # Предрасчитанные блоки для других срезов (справа от основной таблицы)
    block_col: int = last_col + 3
    ws.cell(row=4, column=block_col, value="Другие срезы (предрасчёт)")
    ws.cell(row=4, column=block_col).font = Font(bold=True)
    cursor_row: int = 5
    shown: int = 0
    for tb in tb_values[:3]:
        for metric in metrics:
            for indicator in indicators[:3]:
                if tb == default_tb and metric == default_metric and indicator == default_indicator:
                    continue
                sub_matrix: dict[str, Any] = build_pivot_matrix(pivot_flat, tb, metric, indicator, config)
                if not sub_matrix.get("rows") and not sub_matrix.get("products"):
                    continue
                title: str = f"{tb} | {metric_labels.get(metric, metric)} | {indicator_labels.get(indicator, indicator)}"
                ws.cell(row=cursor_row, column=block_col, value=title)
                cursor_row += 1
                end_row, _ = _write_matrix_table(
                    ws, sub_matrix, start_row=cursor_row, start_col=block_col, row_header=row_header
                )
                cursor_row = end_row + 2
                shown += 1
                if shown >= 4:
                    break
            if shown >= 4:
                break
        if shown >= 4:
            break


def write_charts_sheet(
    wb,
    distribution_series: list[dict[str, Any]],
    config: dict[str, Any],
    sheet_name: str,
) -> None:
    """Лист с графиками кумулятивных кривых (лиды × дни)."""
    ws = wb.create_sheet(sheet_name)
    dash_cfg: dict[str, Any] = config.get("dashboard", {})
    default_tb: str = str(dash_cfg.get("default_tb", dash_cfg.get("all_tb_label", "__ALL__")))
    default_metric: str = str(dash_cfg.get("default_metric", "days_on_stage"))
    max_series: int = int(dash_cfg.get("excel_max_chart_series", 8))

    filtered: list[dict[str, Any]] = [
        s
        for s in distribution_series
        if str(s.get("tb")) == default_tb and str(s.get("metric")) == default_metric
    ]
    filtered.sort(key=lambda item: (-int(item.get("total_leads", 0)), str(item.get("product"))))
    filtered = filtered[:max_series]

    if not filtered:
        ws["A1"] = "Нет данных для графика с выбранными фильтрами по умолчанию."
        return

    ws["A1"] = f"Графики: ТБ={default_tb}, метрика={default_metric}"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = "Ось X — число лидов (накоп.), ось Y — срок в днях. Полный интерактив — HTML/дашборд."
    ws["A2"].font = Font(italic=True)

    chart_row: int = 4
    chart_col_offset: int = 0
    charts_per_row: int = 2

    for idx, series in enumerate(filtered):
        points: list[dict[str, int]] = series_chart_points(series)
        if not points:
            continue

        start_row: int = chart_row + (idx // charts_per_row) * 22
        start_col: int = 1 + (idx % charts_per_row) * 10

        title: str = f"{series.get('row_key') or series.get('product')} | {series.get('stage_key')}"
        ws.cell(row=start_row, column=start_col, value=title)
        ws.cell(row=start_row, column=start_col).font = Font(bold=True)

        data_start: int = start_row + 1
        ws.cell(row=data_start, column=start_col, value="Лид №")
        ws.cell(row=data_start, column=start_col + 1, value="Дней")
        for offset, point in enumerate(points, start=1):
            ws.cell(row=data_start + offset, column=start_col, value=point["lead_index"])
            ws.cell(row=data_start + offset, column=start_col + 1, value=point["days"])

        data_end: int = data_start + len(points)
        chart: ScatterChart = ScatterChart()
        chart.title = title
        chart.style = 2
        chart.x_axis.title = "Число лидов"
        chart.y_axis.title = "Дней"
        chart.width = 16
        chart.height = 10

        xvalues = Reference(ws, min_col=start_col, min_row=data_start + 1, max_row=data_end)
        yvalues = Reference(ws, min_col=start_col + 1, min_row=data_start + 1, max_row=data_end)
        ser = Series(yvalues, xvalues, title="Срок")
        ser.marker = Marker(symbol="circle", size=5)
        ser.graphicalProperties.line.width = 18000
        chart.series.append(ser)

        anchor_col: int = start_col + 3
        anchor_row: int = start_row
        ws.add_chart(chart, f"{get_column_letter(anchor_col)}{anchor_row}")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10


def add_visualization_sheets(
    wb,
    pivot_flat: list[dict[str, Any]],
    distribution_series: list[dict[str, Any]],
    config: dict[str, Any],
) -> None:
    """Добавляет скрытые данные, матрицу и графики."""
    out_cfg: dict[str, Any] = config.get("output", {})
    sheet_names: dict[str, str] = out_cfg.get("excel_sheets", {})

    flat_ws = wb.create_sheet("_pivot_flat")
    write_hidden_pivot_flat(flat_ws, pivot_flat)

    chart_ws = wb.create_sheet("_chart_src")
    write_hidden_chart_source(chart_ws, distribution_series)

    matrix_name: str = sheet_names.get("matrix", "Матрица")
    matrix_ws = wb.create_sheet(matrix_name)
    write_matrix_sheet(matrix_ws, pivot_flat, distribution_series, config)

    charts_name: str = sheet_names.get("charts", "Графики")
    write_charts_sheet(wb, distribution_series, config, charts_name)

    logger.info("Добавлены листы визуализации: %s, %s", matrix_name, charts_name)
